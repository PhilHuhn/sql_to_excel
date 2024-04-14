import pandas as pd
from sqlalchemy import create_engine

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import Color
from openpyxl.drawing.image import Image
import json

connection_string = 'sqlite:///emission_data.db'


class ExcelGenerator:
    def __init__(self, json_file: str, test_mode: bool = False):
        self.format_data = {}
        self.json_file = json_file

        self.query_results = {}
        self.test_mode = test_mode
        self.wb = Workbook()
        self.colors = {
            'black': '000000',
            'white': 'FFFFFF',
            'red': 'FF0000',
            'green': '00FF00',
            'blue': '0000FF',
            'yellow': 'FFFF00',
            'cyan': '00FFFF',
            'magenta': 'FF00FF'
        }

    @property
    def json_file(self):
        return self._json_file

    @json_file.setter
    def json_file(self, value):
        self._json_file = value

        with open(self._json_file, 'r') as f:
            self.format_data = json.load(f)

    def query_data(self):
        """
        Queries the data from the database and stores it in the query_results dictionary.
        if test_mode is True, it reads the data from a csv file.
        """
        if self.test_mode:
            self.query_results["emission_data"] = pd.read_csv('emission_data.csv')
        else:
            engine = create_engine(connection_string)
            for query in self.format_data['queries']:
                self.query_results[query['name']] = pd.read_sql(query['query'], engine)

    def set_colors(self, colors: dict = None):
        """
        Reads the colors from the format_data dictionary and updates the colors dictionary.
        """
        if colors is None:
            self.colors.update(self.format_data.get('colors', {}))
        else:
            self.colors.update(colors)
        self.colors = {k: v.replace('#', '') for k, v in self.colors.items()}

    def format_range(self, sheet: Worksheet, formatting: dict):
        """
        Formats a range of cells based on the formatting dictionary.

        :param sheet: Worksheet object
        :param formatting: dictionary with formatting information
        """
        # Image formatting
        if 'image' in formatting:
            self.format_image(sheet, formatting)
        # column and row formatting
        elif ":" in formatting['start'] and formatting['start'] == formatting.get('end', formatting['start']):
            self.format_column(sheet, formatting['start'], formatting)
        # Single cell formatting
        elif formatting['start'] == formatting.get('end', formatting['start']):
            cell = sheet[formatting['start']]
            self.format_cell(cell, formatting)
        # Range formatting
        else:
            for row in sheet[formatting['start']:formatting['end']]:
                for cell in row:
                    self.format_cell(cell, formatting)

    def format_image(self, sheet: Worksheet, formatting: dict):
        """
        Inserts an image into a cell based on the image_data dictionary.

        :param sheet: Worksheet object
        :param formatting: dictionary with image information
        """
        image = Image(formatting['image']['file'])
        sheet.add_image(image, formatting['image']['start'])

    def format_column(self, sheet: Worksheet, start: str, formatting: dict):
        """
        Formats a column based on the formatting dictionary.

        :param sheet: Worksheet object
        :param start: start cell of the column
        :param formatting: dictionary with formatting information
        """
        start_coordinates = start.split(':')

        sheet.column_dimensions[start_coordinates[0]].width = formatting.get('width', 10)
        if start_coordinates[0] == start_coordinates[1]:
            for cell in sheet[start_coordinates[0]]:
                self.format_cell(cell, formatting)
        else:
            for cell in sheet[start_coordinates[0]:start_coordinates[1]]:
                self.format_cell(cell, formatting)

    def format_cell(self, cell: Cell, formatting: dict):
        """
        Formats a cell based on the formatting dictionary.

        :param cell: Cell object
        :param formatting: dictionary with formatting information
        """
        if 'background_color' in formatting:
            cell.fill = PatternFill(
                start_color=self.get_color(formatting['background_color']),
                fill_type="solid"
            )
        if 'text' in formatting:
            cell.font = Font(
                color=self.get_color(formatting['text'].get('font_color')),
                size=formatting['text'].get('font_size'),
                bold=formatting['text'].get('bold'),
                italic=formatting['text'].get('italic'),
                underline=formatting['text'].get('underline')
            )
            if 'number_format' in formatting["text"]:
                cell.number_format = formatting['text']['number_format']
        if 'value' in formatting:
            cell.value = formatting['value']
        if 'data' in formatting:
            cell.value = self.process_formula(formatting['data'])

    def process_formula(self, data: dict):
        """
        Processes a formula and returns the result.

        :param data: dictionary with query and formula information
        :return: result of the formula
        """
        query = data['query']
        formula = data['formula']

        df = self.query_results[query]

        column = formula['column']
        function = formula['function']

        result = df.agg({column: function})
        return result[column]

    def get_color(self, color: str = None) -> Color:
        """
        Returns the color object for the given color name or hex code.

        :param color: color name or hex code
        :return: openpyxl color object
        """
        if color is None:
            return None
        else:
            return Color(rgb=self.colors.get(color, color))

    def format_data_table(self, sheet: Worksheet, formatting: dict):
        """
        Puts a pandas dataframe in a designated worksheet with formatting and column names.
        Column names are provided in the formatting dictionary with the key 'columns'.
        Location of the top left cell is provided in the formatting dictionary with the key 'start'.

        :param sheet: Worksheet object
        :param formatting: dictionary with formatting information
        """
        # Reorder the columns to match the order specified in the formatting and write the DataFrame to the worksheet
        df = self.query_results[formatting['data']['query']]

        first_cell = sheet[formatting.get('start', 'A1')]
        # Format header row and cells in the column
        for j, col in enumerate(formatting['columns']):
            header_cell = sheet.cell(row=first_cell.row, column=first_cell.column + j)
            header_cell.value = col['display_name']
            self.format_cell(header_cell, col)

            # Write column data to the worksheet
            for i in range(1, len(df) + 1):
                cell = sheet.cell(row=first_cell.row + i, column=first_cell.column + j)
                cell.value = df.iloc[i - 1][col['sql_name']]

                # Format cell
                if 'column_format' in col:
                    self.format_cell(cell, col['column_format'])

    def generate_excel(self, output_file: str):
        """
        Generates an Excel file based on formatting information in a json file.

        :param output_file: file name of the output Excel file
        """
        self.query_data()
        self.set_colors()

        for sheet_data in self.format_data['sheets']:
            sheet = self.wb.create_sheet(title=sheet_data['sheet_name'])
            for formatting in sheet_data['formatting']:
                if formatting['type'] == 'range':
                    self.format_range(sheet, formatting)
                elif formatting['type'] == 'table':
                    self.format_data_table(sheet, formatting)
        # drop the first sheet
        self.wb.remove(self.wb.worksheets[0])
        self.wb.save(output_file)
